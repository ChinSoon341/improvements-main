import os
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from datetime import datetime
import time
import asyncio
import pandas as pd

# ANSI escape codes for text colors
COLOR_RED = '\033[91m'
COLOR_GREEN = '\033[92m'
COLOR_YELLOW = '\033[93m'
COLOR_BLUE = '\033[94m'
COLOR_RESET = '\033[0m'


def append_to_excel(existing_excel_path, new_data):
    try:
        # Read existing data from the Excel file if it exists
        if os.path.exists(existing_excel_path):
            existing_df = pd.read_excel(existing_excel_path)

            # Concatenate existing data with new data
            updated_df = pd.concat([existing_df, new_data], ignore_index=True)

            # Remove duplicates based on all columns
            updated_df.drop_duplicates(inplace=True)

            # Save the updated DataFrame to the Excel file
            updated_df.to_excel(existing_excel_path, index=False)

        else:
            # If the file doesn't exist, create a new DataFrame with the new data
            new_data.to_excel(existing_excel_path, index=False)

        print(f"{COLOR_GREEN}Data appended to: {existing_excel_path}{COLOR_RESET}")

    except Exception as e:
        print(f"{COLOR_RED}An error occurred while appending data: {e}{COLOR_RESET}")



def initialize_driver(webdriver_path):
    chrome_options = webdriver.ChromeOptions()
    chrome_options.add_argument(f'--webdriver={webdriver_path}')

    #chrome_options.add_argument('--headless') #runin background
    return webdriver.Chrome(options=chrome_options)


async def search_and_display(search_term, driver, existing_df, output_excel_path):
    url = 'https://www.gebiz.gov.sg/'
    driver.get(url)

    # Wait for a few seconds to let the page load (you might need to adjust the wait time)
    time.sleep(0.2)

    checkbox_title = driver.find_element(By.ID, 'contentForm:j_id55_0')
    checkbox_title.click()

    # Find the search input field and submit a query
    search_input = driver.find_element(By.CSS_SELECTOR, 'input[name="contentForm:searchBar_searchBar_INPUT-SEARCH"]')
    search_input.send_keys(search_term)
    search_input.send_keys(Keys.RETURN)

    wait = WebDriverWait(driver, 10)

    try:
        try:
            status_element = wait.until(EC.presence_of_element_located((
                    By.XPATH, 
                "//div[contains(@class, 'formSectionHeader6_TEXT') and contains(text(), 'Tender -') and //div[@class='formOutputText_VALUE-DIV ' and @style='text-align: left;']]"
            )))
        except TimeoutException:
            status_element = wait.until(EC.presence_of_element_located((
                    By.XPATH, 
                "//div[contains(@class, 'formSectionHeader6_TEXT') and contains(text(), 'Quotation -') and //div[@class='formOutputText_VALUE-DIV ' and @style='text-align: left;']]"
            )))

        
        # Using XPath to locate the additional elements
        status_element22 = wait.until(EC.presence_of_all_elements_located(
           (By.XPATH, './/a[@class="commandLink_TITLE-BLUE"]')
        ))

            
        # Using XPath to locate the additional elements
        additional_elements = wait.until(EC.presence_of_all_elements_located((
            By.XPATH,
            "//div[@class='formOutputText_VALUE-DIV ' and @style='text-align: left;']"
        )))

        # Extract text from the elements
        tender_number = status_element.text.strip().replace("\n", " ")

                # Extract text from the elements
        tender_name_list = [element.text.strip().replace("\n", " ") for element in status_element22]

        # For demonstration purposes, joining the list into a single string
        tender_name = ', '.join(tender_name_list)

        # Capture the existing values of 'Agency', 'Published', and 'Procurement Category'
        agency = additional_elements[0].text.strip().replace("\n", " ")
        published = additional_elements[1].text.strip().replace("\n", " ")
        procurement_category = additional_elements[2].text.strip().replace("\n", " ")

        # Using XPath to locate the additional elements for Awarded To, Award Value, and Award Date
        subtitle_elements = wait.until(EC.presence_of_all_elements_located((
            By.XPATH,
            "//div[@class='formOutputText_HIDDEN-LABEL outputText_SUBTITLE-BLACK' and @style='text-align: left;']"
        )))

        # Initialize variables for Awarded To, Award Value, and Award Date
        awarded_to = None
        award_value = None
        
        # Loop through subtitle elements to dynamically assign values to columns
        for index, subtitle_element in enumerate(subtitle_elements):
            text = subtitle_element.text.strip()

            if index == 0:
                awarded_to = text
            elif index == 1:
                award_value = text
                placeholder = text
        
        checkbox_title2 = driver.find_element(By.XPATH, "//a[@class='commandLink_ACTION-BLUE' and text()='Multiple Suppliers']")
        checkbox_title2.click()

        # Using WebDriverWait to wait for the presence of the div elements with common structure
        div_elements = wait.until(EC.presence_of_all_elements_located((By.XPATH, "//div[contains(@class, 'formOutputText_HIDDEN-LABEL') and contains(@class, 'outputText_TITLE-BLACK')]")))

        multipleSup = []

        with open(r'C:\Users\ADAM7\Desktop\Projects_demo\04_gebizReport\temp\multiple_Supplier_extract.txt', 'w', encoding='utf-8') as output_file:
            for div_element in div_elements:
                # Extract the text content of each matching element
                div_text = div_element.text

                if "PTE. LTD." in div_text or "PTE LTD" in div_text or "PRIVATE LIMITED" in div_text or "pte ltd" in div_text or "pte. ltd." in div_text:
                    multipleSup.append(div_text)

                output_file.write(f"{div_text}\n")

        # Create a DataFrame with the new columns and original values
        data = {
            'Tender Number': [tender_number],
            'Description' : [tender_name],
            'Agency': [agency],
            'Published': [published],
            'Procurement Category': [procurement_category],
            'Awarded To' : [', '.join(multipleSup)],
            'Award Value': [awarded_to],
            'Awarded Date': [award_value]
        }
        new_df = pd.DataFrame(data)

        # Append the new data to the existing DataFrame
        append_to_excel(output_excel_path, new_df)

        print(f"{COLOR_GREEN}Extraction complete. Data appended to: {output_excel_path}{COLOR_RESET}")

        time.sleep(0.5)

    except Exception as e:
        print(f"{COLOR_RED}An error occurred: {e}{COLOR_RESET}")


async def process_tender_numbers(driver, tender_numbers, existing_df, output_excel_path):
    for tender_number in tender_numbers:
        print(f"Processing Tender Number: {tender_number}")
        await search_and_display(tender_number, driver, existing_df, output_excel_path)


def start_menu():
    webdriver_path = r'C:\Users\chinsoont.BECKHOFF\Downloads'
    driver = initialize_driver(webdriver_path)

    try:
        # Read the Excel file with the list of tender numbers
        excel_path = r'C:\Users\ADAM7\Desktop\Projects_demo\04_gebizReport\temp\exc\flags\f_mulSup.xlsx'
        existing_df = pd.read_excel(excel_path)

        # Define the output_excel_path here
        output_excel_path = r'C:\Users\ADAM7\Desktop\Projects_demo\04_gebizReport\temp\exc\r2_fixed.xlsx'

        # Extract tender numbers from the DataFrame
        tender_numbers = existing_df['Multiple Suppliers'].astype(str).tolist()

        # Run the asyncio event loop to process tender numbers
        asyncio.run(process_tender_numbers(driver, tender_numbers, existing_df, output_excel_path))

    except Exception as e:
        print(f"{COLOR_RED}An error occurred: {e}{COLOR_RESET}")
    finally:
        driver.quit()

if __name__ == "__main__":
    start_menu()


# merged back to the r1_removed and create new output
    
def add_entries(input_file, output_file):
    # Read entries from the input file
    input_df = pd.read_excel(input_file)

    # Read entries from the output file if it exists, otherwise create an empty DataFrame
    if os.path.exists(output_file):
        output_df = pd.read_excel(output_file)
    else:
        output_df = pd.DataFrame()

    # Add entries from input file to the output DataFrame
    output_df = output_df.append(input_df, ignore_index=True)

    # Write the combined DataFrame to the output file
    output_df.to_excel(output_file, index=False)

    # Apply formatting to the output file
    apply_formatting(output_file)

def apply_formatting(output_file):
    # Load the workbook
    workbook = load_workbook(output_file)
    worksheet = workbook.active

    # Apply auto width to all columns
    for column_cells in worksheet.columns:
        max_length = 0
        column = column_cells[0].column_letter
        for cell in column_cells:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        worksheet.column_dimensions[column].width = adjusted_width

    # Apply bold font to headers
    bold_font = Font(bold=True)
    for cell in worksheet[1]:
        cell.font = bold_font

    # Apply grey fill to header cells
    grey_fill = PatternFill(start_color="C0C0C0", end_color="C0C0C0", fill_type="solid")
    for cell in worksheet[1]:
        cell.fill = grey_fill

    # Save the workbook
    workbook.save(output_file)

if __name__ == "__main__":
    input_file_path = r'C:\Users\ADAM7\Desktop\Projects_demo\04_gebizReport\temp\exc\r2_fixed.xlsx'
    output_directory = r'C:\Users\ADAM7\Desktop\Projects_demo\04_gebizReport'

    # Generate output filename based on current date and time
    current_datetime = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    output_filename = f"{current_datetime}_geBIZ_Report.xlsx"
    output_file_path = os.path.join(output_directory, output_filename)

    # Existing file to which entries need to be added
    existing_output_file_path = os.path.join(output_directory, 'temp', 'exc', 'r1_main_removed.xlsx')

    add_entries(input_file_path, existing_output_file_path)

    print("Entries added successfully to 'r1_main_removed.xlsx'")

    # Now rename the existing output file to the desired name
    os.rename(existing_output_file_path, output_file_path)

    print("Output saved as:", output_file_path)



