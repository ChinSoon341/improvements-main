import os
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import time
import asyncio
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from telegram import Bot
from telegram.ext import Updater

# ANSI escape codes for text colors
COLOR_RED = '\033[91m'
COLOR_GREEN = '\033[92m'
COLOR_YELLOW = '\033[93m'
COLOR_BLUE = '\033[94m'
COLOR_RESET = '\033[0m'

# Path to the log Excel file
LOG_FILE_PATH = r'C:\Users\ADAM7\Desktop\Projects_demo\02_telegrambot\logs\tenderMointorLog.xlsx'

# Specify the path to your webdriver (replace with the actual path to your chromedriver.exe)
webdriver_path = r'C:\Users\chinsoont.BECKHOFF\Downloads'

# Telegram bot token
TELEGRAM_BOT_TOKEN = '6540947414:AAGVoA9X-vw-Uq1b1YqLvTH8YkQX1UlfNkU'
DEFAULT_CHAT_ID = '1471225821'  # Replace with your default chat ID
OTHER_CHAT_ID = '1471225821'  # Replace with another chat ID

async def notify_on_telegram(search_term, status_text, awarded_to, award_value, award_date, recipient_chat_id=None):
    bot = Bot(token=TELEGRAM_BOT_TOKEN)

    # If recipient_chat_id is None, use the default chat ID
    chat_id = recipient_chat_id if recipient_chat_id else DEFAULT_CHAT_ID

    # Message formatting for better clarity (e.g., bolding the status_text)
    message = f"Status update for tender {search_term}:\n\n<b>{status_text}</b>\n\nAwarded to: {awarded_to}\nAward Value: {award_value}\nAward Date: {award_date}\n\nLast Updated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"

    # Use parse_mode='HTML' to enable text formatting (e.g., bold)
    bot.send_message(chat_id=chat_id, text=message, parse_mode='HTML')

async def search_and_display(search_term):
    chrome_options = webdriver.ChromeOptions()
    chrome_options.add_argument(f'--webdriver={webdriver_path}')

    driver = webdriver.Chrome(options=chrome_options)
    url = 'https://www.gebiz.gov.sg/'
    driver.get(url)
    time.sleep(0.2)

    checkbox_title = driver.find_element(By.ID, 'contentForm:j_id55_0')
    checkbox_title.click()

    search_input = driver.find_element(By.CSS_SELECTOR, 'input[name="contentForm:searchBar_searchBar_INPUT-SEARCH"]')
    search_input.send_keys(search_term)
    search_input.send_keys(Keys.RETURN)
    time.sleep(0.5)

    wait = WebDriverWait(driver, 10)

    try:
        status_element = wait.until(EC.presence_of_element_located((
            By.XPATH, 
            '//div[@class="label_MAIN label_WHITE-ON-GRAY" or '
            '(@class="label_MAIN label_WHITE-ON-GREEN" and text()="AWARDED") or '
            '(@class="label_MAIN label_WHITE-ON-GRAY" and text()="OPEN") or '
            '(@class="label_MAIN label_WHITE-ON-LIGHT-GRAY" and (text()="CLOSED" or '
            'text()="CANCELLED" or text()="NO AWARD"))]'
        )))

        status_text = status_element.text.strip()
        status_text = status_text.replace("\n", " ")

        subtitle_elements = wait.until(EC.presence_of_all_elements_located((
            By.XPATH,
            "//div[@class='formOutputText_HIDDEN-LABEL outputText_SUBTITLE-BLACK' and @style='text-align: left;']"
        )))

        awarded_to = None
        award_value = None
        award_date = None

        for index, subtitle_element in enumerate(subtitle_elements):
            text = subtitle_element.text.strip()

            if index == 0:
                awarded_to = text
            elif index == 1:
                award_value = text
            elif index == 2:
                award_date = text

        log_to_excel(search_term, status_text, awarded_to, award_value, award_date)

    except Exception as e:
        print(f"{COLOR_RED}An error occurred: {e}{COLOR_RESET}")

    driver.quit()

def log_to_excel(search_term, status_text, awarded_to, award_value, award_date):
    if not os.path.exists(LOG_FILE_PATH):
        wb = Workbook()
        ws = wb.active
        ws.append(["Search Term", "Status", "Awarded To", "Award Value", "Award Date", "Date and Time", "Chat ID"])
    else:
        wb = load_workbook(LOG_FILE_PATH)
        ws = wb.active

    # Formatting the header row
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color='C0C0C0', end_color='C0C0C0', fill_type='solid')

    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill

    chat_id = DEFAULT_CHAT_ID  # You can change this according to your logic
    current_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    ws.append([search_term, status_text, awarded_to, award_value, award_date, current_time, chat_id])

    # Auto-width for all columns
    for column in ws.columns:
        max_length = 0

        # Including the header row in the calculation
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass

        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column[0].column_letter].width = adjusted_width

    wb.save(LOG_FILE_PATH)
            
def start_menu():
    with open(r'C:\Users\ADAM7\Desktop\Projects_demo\02_telegrambot\app_data\tenderNo.txt', 'r') as file:
        file_content = file.read()
    asyncio.run(search_and_display(file_content))

if __name__ == "__main__":
    start_menu()
