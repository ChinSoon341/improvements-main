import os
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import time
import openpyxl
from datetime import datetime

# ANSI escape codes for text colors
COLOR_RED = '\033[91m'
COLOR_GREEN = '\033[92m'
COLOR_YELLOW = '\033[93m'
COLOR_BLUE = '\033[94m'
COLOR_RESET = '\033[0m'

# Specify the path to your OneDrive folder
ONE_DRIVE_FOLDER = r'C:\Users\chinsoont.BECKHOFF\OneDrive - Singapore Polytechnic'

# Specify the name of the Excel file
EXCEL_FILE_NAME = 'iteverysmart.xlsx'

# Create the full path to the Excel file in the OneDrive folder
EXCEL_FILE_PATH = os.path.join(ONE_DRIVE_FOLDER, EXCEL_FILE_NAME)

def clear_screen():
    os.system('cls' if os.name == 'nt' else 'clear')

def save_search_term(search_term):
    with open("search_history.txt", "w") as file:
        file.write(search_term)
    return search_term

def load_last_search_term():
    try:
        with open("search_history.txt", "r") as file:
            return file.read().strip()
    except FileNotFoundError:
        return ""

def save_to_excel(search_term, status_text):
    # Check if the Excel file exists
    if not os.path.exists(EXCEL_FILE_PATH):
        # Create a new Excel file with headers if it doesn't exist
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Tender Number", "Status", "Time Updated"])
        wb.save(EXCEL_FILE_PATH)
    else:
        # Load the existing workbook
        wb = openpyxl.load_workbook(EXCEL_FILE_PATH)
        ws = wb.active

    # Check if the tender number is already in the sheet
    for row in ws.iter_rows(min_row=2, max_col=1, max_row=ws.max_row):
        for cell in row:
            if cell.value == search_term:
                # If the tender number is found, update the status and time
                ws.cell(row=cell.row, column=2, value=status_text)
                ws.cell(row=cell.row, column=3, value=datetime.now().strftime('%Y-%m-%d %H:%M:%S'))
                wb.save(EXCEL_FILE_PATH)
                return

    # If the tender number is not found, append a new row
    ws.append([search_term, status_text, datetime.now().strftime('%Y-%m-%d %H:%M:%S')])
    wb.save(EXCEL_FILE_PATH)

    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column = [cell for cell in column]
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 4)
        ws.column_dimensions[column[0].column_letter].width = adjusted_width

    # Styling Headers
    for header_cell in ws[1]:
        header_cell.font = openpyxl.styles.Font(bold=True, color='000000')  # Black font color
        header_cell.fill = openpyxl.styles.PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")  # Light gray background

    # Conditional Formatting for Status
    
    orange_fill = openpyxl.styles.PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")  # Orange fill
    red_fill = openpyxl.styles.PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")  # Red fill for "CLOSED"
    green_fill = openpyxl.styles.PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")  # Green fill for "OPEN"

    for row in ws.iter_rows(min_row=2, max_col=2, max_row=ws.max_row):
        for cell in row:
            if cell.value == "PENDING AWARD":
                cell.fill = orange_fill
            elif cell.value == "CLOSED":
                cell.fill = red_fill
            elif cell.value == "OPEN":
                cell.fill = green_fill

    wb.save(EXCEL_FILE_PATH)

def search_and_display(search_term):
    clear_screen()
    # Set the path to your webdriver (replace with the actual path to your chromedriver.exe)
    webdriver_path = r'C:\Users\chinsoont.BECKHOFF\Downloads'

    # Set the webdriver path using the executable_path argument
    chrome_options = webdriver.ChromeOptions()
    chrome_options.add_argument(f'--webdriver={webdriver_path}')

    # Add the following line to run Chrome in headless mode
    chrome_options.add_argument('--headless')

    driver = webdriver.Chrome(options=chrome_options)

    url = 'https://www.gebiz.gov.sg/'

    # Open the website in the browser
    driver.get(url)

    # Wait for a few seconds to let the page load (you might need to adjust the wait time)
    time.sleep(0.2)

    checkbox_title = driver.find_element(By.ID, 'contentForm:j_id55_0')
    checkbox_title.click()

    # Find the search input field and submit a query
    search_input = driver.find_element(By.CSS_SELECTOR, 'input[name="contentForm:searchBar_searchBar_INPUT-SEARCH"]')
    search_input.send_keys(search_term)
    search_input.send_keys(Keys.RETURN)

    # Wait for a few seconds to let the search results load (you might need to adjust the wait time)
    time.sleep(0.5)

    wait = WebDriverWait(driver, 10)

    try:
        # Using XPath to locate the status element based on its class
        status_element = wait.until(EC.presence_of_element_located((By.XPATH, '//div[@class="label_MAIN label_WHITE-ON-GRAY"]')))

        # Using XPath to locate the document title element based on its class
        title_element = wait.until(EC.presence_of_element_located((By.XPATH, '//a[@class="commandLink_TITLE-BLUE"]')))

        # Using XPath to locate the category element based on its position
        category_element = wait.until(EC.presence_of_element_located((By.XPATH, '(//div[contains(@class, "formOutputText_VALUE-DIV")])[3]')))

        # Using XPath to locate the category element based on its position
        pub_element = wait.until(EC.presence_of_element_located((By.XPATH, '(//div[contains(@class, "formOutputText_VALUE-DIV")])[2]')))

        # Check if all elements are found
        status_text = status_element.text.strip()
        title_text = title_element.text.strip()
        category_text = category_element.text.strip()
        pub_time = pub_element.text.strip()

        # Replace "\n" with a space in the output
        status_text = status_text.replace("\n", " ")
        title_text = title_text.replace("\n", " ")
        category_text = category_text.replace("\n", " ")
        pub_time = pub_time.replace("\n", " ")

        # Display the information in a more structured format with colors
        clear_screen()
        print(f"\n{COLOR_RESET}Tender Number:{COLOR_GREEN} {search_term} {COLOR_RESET}")
        print(f"Title: {title_text}")
        print(f"Category:  {category_text}")
        print(f"Published: {pub_time}")
        print(f"Status: {COLOR_YELLOW} {status_text} {COLOR_RESET}")
        print(f"Time Updated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')} {COLOR_RESET}")

        # Update the search history immediately
        save_search_term(search_term)

        # Save to Excel
        save_to_excel(search_term, status_text)

    except Exception as e:
        print(f"{COLOR_RED}An error occurred: {e}{COLOR_RESET}")

    # Prompt to exit back to the start menu
    input(f"{COLOR_RESET}\nPress Enter to go back to the start menu...")

    # Close the browser
    driver.quit()

def start_menu():
    last_search_term = load_last_search_term()

    while True:
        clear_screen()
        print(f"{COLOR_YELLOW}\n### geBIZ Search ###")
        print(f"1. Search")
        print(f"2. Exit\n")

        choice = input(f"{COLOR_YELLOW}Enter your choice (1 or 2): ")

        if not choice or choice == "1":
            clear_screen()
            if last_search_term:
                use_last_term = input(f"{COLOR_YELLOW}Do you want to use the last search term {COLOR_GREEN}'{last_search_term}'{COLOR_RESET}? (y/n): {COLOR_RESET}")
                if not use_last_term or use_last_term.lower() == 'y':
                    search_term = last_search_term
                else:
                    search_term = input(f"{COLOR_YELLOW}Enter the new search term: {COLOR_RESET}")
            else:
                search_term = input(f"{COLOR_YELLOW}Enter the search term: {COLOR_RESET}")

            last_search_term = save_search_term(search_term)  # Update last_search_term
            search_and_display(search_term)

        elif choice == "2":
            print(f"{COLOR_RESET}Exiting the program. Goodbye!{COLOR_RESET}")
            break
        else:
            print(f"{COLOR_RED}Invalid choice. Please enter 1 or 2.{COLOR_RESET}")

if __name__ == "__main__":
    start_menu()
