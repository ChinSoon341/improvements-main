import os
import azure.ai.vision as sdk
import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

# Exported to environment variable, safer and more secure for sensitive information like APIs
VISION_ENDPOINT = os.environ.get("VISION_ENDPOINT") 
VISION_KEY = os.environ.get("VISION_KEY")

#to set: setx VISION_KEY 

if VISION_ENDPOINT is None or VISION_KEY is None:
    raise ValueError("VISION_ENDPOINT and VISION_KEY must be set in the environment")

service_options = sdk.VisionServiceOptions(VISION_ENDPOINT, VISION_KEY)

# Specify the paths to your local image files
image_paths = [
    r'C:\Users\ADAM7\Desktop\Projects_demo\03_ocr\output\1_output\01_lead_number.jpg',
    r'C:\Users\ADAM7\Desktop\Projects_demo\03_ocr\output\1_output\02_company.jpg',
    r'C:\Users\ADAM7\Desktop\Projects_demo\03_ocr\output\1_output\03_name.jpg',
    r'C:\Users\ADAM7\Desktop\Projects_demo\03_ocr\output\1_output\07_industry.jpg',
    r'C:\Users\ADAM7\Desktop\Projects_demo\03_ocr\output\1_output\08_interest.jpg',
    r'C:\Users\ADAM7\Desktop\Projects_demo\03_ocr\output\1_output\11_preparedBy.jpg',
    r'C:\Users\ADAM7\Desktop\Projects_demo\03_ocr\output\1_output\06_category_resized.jpg\06_category_resized.jpg',
    r'C:\Users\ADAM7\Desktop\Projects_demo\03_ocr\output\1_output\09_product_resized.jpg\09_product_resized.jpg',
    r'C:\Users\ADAM7\Desktop\Projects_demo\03_ocr\output\1_output\05_nameCards.jpg'
]

# Specify corresponding headers
headers = ["Lead Number", "Company", "Name", "Industry", "Interest", "Prepared by","Category","Product", "Etc"]

# Initialize the Excel workbook and sheet
excel_file_path = r'C:\Users\ADAM7\Desktop\Projects_demo\03_ocr\test.xlsx'

# Load existing workbook if exists, otherwise create a new one
if os.path.exists(excel_file_path):
    wb = openpyxl.load_workbook(excel_file_path)
else:
    wb = openpyxl.Workbook()

# Select the active sheet
sheet = wb.active

# Write headers to the Excel sheet if it's a new file
if not sheet.rows:
    sheet.append(headers)
    # Header formatting
    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")

# Initialize a list to store the extracted information
extracted_info_list = []

# Iterate through each image path and extract information using OCR
for image_path, header in zip(image_paths, headers):
    with open(image_path, 'rb') as image_file:
        image_buffer = image_file.read()

    image_source_buffer = sdk.ImageSourceBuffer()
    image_source_buffer.image_writer.write(image_buffer)
    vision_source = sdk.VisionSource(image_source_buffer=image_source_buffer)

    analysis_options = sdk.ImageAnalysisOptions()
    analysis_options.features = (
        sdk.ImageAnalysisFeature.TEXT
    )
    analysis_options.language = "en"
    analysis_options.gender_neutral_caption = False

    image_analyzer = sdk.ImageAnalyzer(service_options, vision_source, analysis_options)
    result = image_analyzer.analyze()

    # Extract relevant information from OCR result
    extracted_info = ""
    if result.reason == sdk.ImageAnalysisResultReason.ANALYZED and result.text is not None:
        for line in result.text.lines:
            extracted_info += line.content + " "

    # Append the extracted information to the list
    extracted_info_list.append(extracted_info)

# Append the list to the Excel sheet as a single row
sheet.append(extracted_info_list)

# Auto-adjust column width
for column in sheet.columns:
    max_length = 0
    column = [cell for cell in column]
    for cell in column:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(cell.value)
        except:
            pass
    adjusted_width = (max_length + 2)
    sheet.column_dimensions[get_column_letter(column[0].column)].width = adjusted_width

# Save the Excel file
wb.save(excel_file_path)

print(f"Completed. Check '{excel_file_path}' for the results.")
