import os
import re
import azure.ai.vision as sdk
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import NamedStyle, Font, Border, Side, PatternFill

# exported to environment variable, safer and more secure for sensitive information like APIs
VISION_ENDPOINT = os.environ.get("VISION_ENDPOINT")
VISION_KEY = os.environ.get("VISION_KEY")

if VISION_ENDPOINT is None or VISION_KEY is None:
    raise ValueError("VISION_ENDPOINT and VISION_KEY must be set in the environment")

service_options = sdk.VisionServiceOptions(VISION_ENDPOINT, VISION_KEY)

# Folder path for the images
folder_path = r'C:\Users\chinsoont.BECKHOFF\OneDrive - Singapore Polytechnic\improvements\03_ocr_raja\v0.2\dataset\trial\png_nodetection\half_restitched'

# Generate a list of image paths dynamically and sort them using a natural sort order
def natural_sort_key(s):
    return [int(text) if text.isdigit() else text.lower() for text in re.split(r'(\d+)', s)]

image_paths = sorted([os.path.join(folder_path, f) for f in os.listdir(folder_path) if f.endswith('.jpg')], key=natural_sort_key)

# List to store the extracted information DataFrames
dfs = []

# Loop through each image in the list
for image_path in image_paths:
    # Read image content and create an image buffer
    with open(image_path, 'rb') as image_file:
        image_buffer = image_file.read()

    image_source_buffer = sdk.ImageSourceBuffer()
    image_source_buffer.image_writer.write(image_buffer)
    vision_source = sdk.VisionSource(image_source_buffer=image_source_buffer)

    analysis_options = sdk.ImageAnalysisOptions()

    analysis_options.features = (
        sdk.ImageAnalysisFeature.CAPTION |
        sdk.ImageAnalysisFeature.TEXT
    )

    analysis_options.language = "en"
    analysis_options.gender_neutral_caption = False

    image_analyzer = sdk.ImageAnalyzer(service_options, vision_source, analysis_options)

    result = image_analyzer.analyze()

    # Create a dictionary to store the extracted information, including image path
    extracted_info = {}

    current_header = None
    current_text = ""
    what_they_want_text = ""

    if result.reason == sdk.ImageAnalysisResultReason.ANALYZED:
        if result.text is not None:
            for line in result.text.lines:
                if line.content.endswith(":"):
                    if current_header is not None:
                        extracted_info[current_header] = current_text.strip()
                        # Combine text for "What they want"
                        if current_header in ["Interest:", "Limitation of current control system:", "Details of current control system (architecture, current brand/model, fieldbus, protocol, etc.):"]:
                            what_they_want_text += current_text.strip() + "\n"
                        # Check for the header "What is the solution they are looking for?" and add it to "What they want"
                        elif current_header == "What is the solution they are looking for?":
                            what_they_want_text += current_text.strip() + "\n"
                    current_header = line.content
                    current_text = ""
                else:
                    current_text += line.content + " "
    
    # Check if the "What is the solution they are looking for?" header was not detected
    if "What is the solution they are looking for?" not in extracted_info:
        # Assume that the text belongs to the previous header (if any)
        what_they_want_text += current_text.strip() + "\n"

    # Add the last entry
    if current_header is not None:
        extracted_info[current_header] = current_text.strip()

    # Combine text for "What they want" and remove the specific header
    what_they_want_text = what_they_want_text.replace("What is the solution they are looking for?", "")
    extracted_info["What they want"] = what_they_want_text.strip()

    # Remove unwanted headers
    for unwanted_header in ["Interest:", "Limitation of currrent control system:", "Details of current control system (architecture, current brand/model, fieldbus, protocol, etc.):", "Category:"
                            ,"Limitation of current control system:"]:
        
        extracted_info.pop(unwanted_header, None)

    # Extract file name from the 'Image Path'
    extracted_info['Image Name'] = os.path.splitext(os.path.basename(image_path))[0]
    extracted_info['Image Path'] = os.path.basename(image_path)  # Include entire path as a separate column

    # Convert the dictionary to a DataFrame and append to the list
    df_image = pd.DataFrame([extracted_info])

    # Remove unwanted columns
    unwanted_columns = ["Image Path"]
    df_image = df_image.drop(columns=unwanted_columns, errors='ignore')

    # Append the modified DataFrame to the list
    dfs.append(df_image)

# Concatenate all DataFrames in the list
df_combined = pd.concat(dfs, ignore_index=True)

# Save the combined DataFrame to an Excel file with styling
output_file_path = r'C:\Users\chinsoont.BECKHOFF\OneDrive - Singapore Polytechnic\improvements\03_ocr_raja\v0.2\dataset\trial\png_nodetection\excel_test_combined_2_images_no_solution_header.xlsx'
with pd.ExcelWriter(output_file_path, engine='openpyxl') as writer:
    df_combined.to_excel(writer, index=False, sheet_name='Sheet1')

    # Access the XlsxWriter workbook and worksheet objects
    workbook = writer.book
    worksheet = writer.sheets['Sheet1']

    # Create a style for headers
    header_style = NamedStyle(name='header_style', font=Font(bold=True), border=Border(bottom=Side(border_style='thin')), fill=PatternFill(start_color="C0C0C0", end_color="C0C0C0", fill_type="solid"))

    # Apply the style to headers
    for cell in worksheet['1:1']:
        cell.style = header_style

    # Set column width to auto
    for column in worksheet.columns:
        max_length = 0
        column = [cell for cell in column]
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        worksheet.column_dimensions[column[0].column_letter].width = adjusted_width

    # Apply borders to all cells
    for row in worksheet.iter_rows():
        for cell in row:
            cell.border = Border(left=Side(border_style='thin'), right=Side(border_style='thin'), top=Side(border_style='thin'), bottom=Side(border_style='thin'))

# Save the workbook
workbook.save(output_file_path)

print(f"Extraction completed. Check '{output_file_path}' for the results.")
