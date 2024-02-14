import os
import azure.ai.vision as sdk
import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

# Load sensitive API keys securely from environment variables
VISION_ENDPOINT = os.environ.get("VISION_ENDPOINT")
VISION_KEY = os.environ.get("VISION_KEY")

if not VISION_ENDPOINT or not VISION_KEY:
    raise ValueError("VISION_ENDPOINT and VISION_KEY must be set in environment variables.")

service_options = sdk.VisionServiceOptions(VISION_ENDPOINT, VISION_KEY)

# Define base directory and folder list
base_dir = r'C:\Users\ADAM7\Desktop\Projects_demo\03_ocr\output'
output_folders = [d for d in os.listdir(base_dir) if os.path.isdir(os.path.join(base_dir, d))]

# Specify fixed headers for consistent output
headers = ["Lead Number", "Company", "Name", "Industry", "Interest", "Prepared by",
            "Category", "Product", "Etc"]

# Use a single Excel file for accumulating extracted data
excel_file_path = r'C:\Users\ADAM7\Desktop\Projects_demo\03_ocr\test.xlsx'
wb = openpyxl.load_workbook(excel_file_path) if os.path.exists(excel_file_path) else openpyxl.Workbook()
sheet = wb.active

# Process each folder sequentially, accumulating data
for output_folder_number in range(1, len(output_folders) + 1):
    image_paths = [  # Construct image paths dynamically for each folder
        rf'C:\Users\ADAM7\Desktop\Projects_demo\03_ocr\output\{output_folder_number}_output\01_lead_number.jpg',
        rf'C:\Users\ADAM7\Desktop\Projects_demo\03_ocr\output\{output_folder_number}_output\02_company.jpg',
        rf'C:\Users\ADAM7\Desktop\Projects_demo\03_ocr\output\{output_folder_number}_output\03_name.jpg',
        rf'C:\Users\ADAM7\Desktop\Projects_demo\03_ocr\output\{output_folder_number}_output\07_industry.jpg',
        rf'C:\Users\ADAM7\Desktop\Projects_demo\03_ocr\output\{output_folder_number}_output\08_interest.jpg',
        rf'C:\Users\ADAM7\Desktop\Projects_demo\03_ocr\output\{output_folder_number}_output\11_preparedBy.jpg',
        rf'C:\Users\ADAM7\Desktop\Projects_demo\03_ocr\output\{output_folder_number}_output\06_category_resized.jpg\06_category_resized.jpg',
        rf'C:\Users\ADAM7\Desktop\Projects_demo\03_ocr\output\{output_folder_number}_output\09_product_resized.jpg\09_product_resized.jpg',
        rf'C:\Users\ADAM7\Desktop\Projects_demo\03_ocr\output\{output_folder_number}_output\05_nameCards.jpg'
    ]

    extracted_info_list = []  # Store extracted info for this folder

    for image_path, header in zip(image_paths, headers):
        with open(image_path, 'rb') as image_file:
            image_buffer = image_file.read()

        image_source_buffer = sdk.ImageSourceBuffer()
        image_source_buffer.image_writer.write(image_buffer)
        vision_source = sdk.VisionSource(image_source_buffer=image_source_buffer)

        analysis_options = sdk.ImageAnalysisOptions()
        analysis_options.features = (
            sdk.ImageAnalysisFeature.TEXT
        )
        analysis_options.language = "en"
        analysis_options.gender_neutral_caption = False

        image_analyzer = sdk.ImageAnalyzer(service_options, vision_source, analysis_options)
        result = image_analyzer.analyze()

        # Extract relevant information from OCR result
        extracted_info = ""
        if result.reason == sdk.ImageAnalysisResultReason.ANALYZED and result.text is not None:
            for line in result.text.lines:
                extracted_info += line.content + " "

        extracted_info_list.append(extracted_info)

    # Append extracted data for this folder to the Excel sheet
    sheet.append(extracted_info_list)

# Auto-adjust column widths based on content
for column in sheet.columns:
    max_length = 0
    for cell in column:
        try:
            max_length = max(max_length, len(str(cell.value)))
        except:
            pass
    sheet.column_dimensions[get_column_letter(column[0].column)].width = max_length + 2

# Save the updated Excel file
wb.save(excel_file_path)

print(f"Completed. Check '{excel_file_path}' for the combined results. New data appended as more folders are processed.")