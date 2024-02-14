import os
import azure.ai.vision as sdk
import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
import pandas as pd
import time

# Load sensitive API keys securely from environment variables
VISION_ENDPOINT = os.environ.get("VISION_ENDPOINT")
VISION_KEY = os.environ.get("VISION_KEY")

if not VISION_ENDPOINT or not VISION_KEY:
    raise ValueError("VISION_ENDPOINT and VISION_KEY must be set in environment variables.")

service_options = sdk.VisionServiceOptions(VISION_ENDPOINT, VISION_KEY)

# Define base directory and folder list
base_dir = r'C:\Users\ADAM7\Desktop\Projects_demo\03_ocr\output'
output_folders = [d for d in os.listdir(base_dir) if os.path.isdir(os.path.join(base_dir, d))]

# Specify fixed headers for consistent output
headers = [ "Company", "Name", "Industry", "Interest",
            "Category", "Product", "Etc"]

# Use a single Excel file for accumulating extracted data
excel_file_path = r'C:\Users\ADAM7\Desktop\Projects_demo\03_ocr\temp\temp1.xlsx'

outputH_excel_path = r'C:\Users\ADAM7\Desktop\Projects_demo\03_ocr\results.xlsx'

wb = openpyxl.load_workbook(excel_file_path) if os.path.exists(excel_file_path) else openpyxl.Workbook()

confidence_wb = openpyxl.Workbook()

sheet = wb.active
confidence_sheet = confidence_wb.active


confidence_excel_file_path = r'C:\Users\ADAM7\Desktop\Projects_demo\03_ocr\logs\confidence_list.xlsx'

# Process each folder sequentially, accumulating data
for output_folder_number in range(1, len(output_folders) + 1):
    image_paths = [  # Construct image paths dynamically for each folder
        rf'C:\Users\ADAM7\Desktop\Projects_demo\03_ocr\output\{output_folder_number}_output\02_company.jpg',
        rf'C:\Users\ADAM7\Desktop\Projects_demo\03_ocr\output\{output_folder_number}_output\03_name.jpg',
        rf'C:\Users\ADAM7\Desktop\Projects_demo\03_ocr\output\{output_folder_number}_output\07_industry.jpg',
        rf'C:\Users\ADAM7\Desktop\Projects_demo\03_ocr\output\{output_folder_number}_output\08_interest.jpg',
        rf'C:\Users\ADAM7\Desktop\Projects_demo\03_ocr\output\{output_folder_number}_output\06_category.jpg',
        rf'C:\Users\ADAM7\Desktop\Projects_demo\03_ocr\output\{output_folder_number}_output\09_product.jpg',
        rf'C:\Users\ADAM7\Desktop\Projects_demo\03_ocr\output\{output_folder_number}_output\05_nameCards.jpg'
    ]

    extracted_info_list = []  # Store extracted info for this folder
    confidence_list = []

    for image_path, header in zip(image_paths, headers):
        with open(image_path, 'rb') as image_file:
            image_buffer = image_file.read()

        time.sleep(0.5)
        image_source_buffer = sdk.ImageSourceBuffer()
        image_source_buffer.image_writer.write(image_buffer)
        vision_source = sdk.VisionSource(image_source_buffer=image_source_buffer)

        analysis_options = sdk.ImageAnalysisOptions()
        analysis_options.features = (
            sdk.ImageAnalysisFeature.TEXT
        )
        analysis_options.language = "en"
        analysis_options.gender_neutral_caption = False

        image_analyzer = sdk.ImageAnalyzer(service_options, vision_source, analysis_options)
        result = image_analyzer.analyze()

        # Extract relevant information from OCR result
        extracted_info = ""
        if result.reason == sdk.ImageAnalysisResultReason.ANALYZED and result.text is not None:
            for line in result.text.lines :
                extracted_info += line.content + " "

                print(f"Word: {line.content}\n")
                for word in line.words:
                    word_content = word.content
                    word_confidence = word.confidence
                    confidence_sheet.append([word_content, word_confidence])

                    if word_confidence <0.90:
                        extracted_info += f"*"

                    print(f"Word: {word.content}\n")
                    print(f"Confidence: {word.confidence}\n")
        time.sleep(0.5)
        extracted_info_list.append(extracted_info)
        
    # Append extracted data for this folder to the Excel sheet
    sheet.append(extracted_info_list)

# Auto-adjust column widths based on content
for column in sheet.columns:
    max_length = 0
    for cell in column:
        try:
            max_length = max(max_length, len(str(cell.value)))
        except:
            pass
    sheet.column_dimensions[get_column_letter(column[0].column)].width = max_length + 2

# Save the updated Excel file
wb.save(excel_file_path)
confidence_wb.save(confidence_excel_file_path)

# Read the Excel file
data = pd.read_excel(excel_file_path, header=None)

# Assign headers to the DataFrame
data.columns = headers

# Save the DataFrame to a new Excel file with auto-adjusted column widths
with pd.ExcelWriter(outputH_excel_path, engine='openpyxl', mode='w') as writer:
    data.to_excel(writer, index=False)

    # Access the openpyxl workbook and worksheet objects
    workbook = writer.book
    worksheet = workbook.active

    # Auto-adjust column width
    for column in worksheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        worksheet.column_dimensions[column_letter].width = adjusted_width


print("Data has been successfully read and saved with headers.")
